from __future__ import annotations

from pathlib import Path
from typing import Optional

from ..types import GridTable, SpanInfo, default_name


def read_excel_table(path: str | Path, *, sheet: Optional[str] = None, table_id: str = "") -> GridTable:
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    grid = [["" for _ in range(max_col)] for _ in range(max_row)]
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                grid[cell.row - 1][cell.column - 1] = str(cell.value).strip()

    span_info: SpanInfo = {}
    seen = set()
    for merged in ws.merged_cells.ranges:
        r1, c1, r2, c2 = merged.min_row, merged.min_col, merged.max_row, merged.max_col
        rowspan, colspan = r2 - r1 + 1, c2 - c1 + 1
        value = grid[r1 - 1][c1 - 1]
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                grid[r - 1][c - 1] = value
        if rowspan > 1 or colspan > 1:
            key = value if value else f"(empty@R{r1}C{c1})"
            if key in seen:
                key = f"{key}@R{r1}C{c1}"
            seen.add(key)
            span_info[key] = {}
            if rowspan > 1:
                span_info[key]["rowspan"] = rowspan
            if colspan > 1:
                span_info[key]["colspan"] = colspan

    wb.close()
    return GridTable.from_grid(
        grid,
        span=span_info,
        table_id=table_id,
        name=default_name(path),
        source_type="excel",
        metadata={"sheet": ws.title},
    )
